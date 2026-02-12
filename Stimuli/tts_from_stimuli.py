#!/usr/bin/env python3
"""Generate English TTS files from an Excel stimulus list."""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import edge_tts
from openpyxl import load_workbook

VOICE_PRESETS: dict[str, list[tuple[str, str]]] = {
    "en_3m3f": [
        ("m1_guy", "en-US-GuyNeural"),
        ("m2_christopher", "en-US-ChristopherNeural"),
        ("m3_ryan", "en-US-BrianNeural"),
        ("f1_aria", "en-US-AriaNeural"),
        ("f2_jenny", "en-US-JennyNeural"),
        ("f3_sonia", "en-US-AnaNeural"),
    ]
}

# Per-voice pronunciation overrides for problematic items.
PRONUNCIATION_OVERRIDES: dict[tuple[str, str], str] = {
    ("m3_ryan", "acorn.mp3"): "A-corn",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate mp3 files from texts in an Excel column."
    )
    parser.add_argument(
        "--xlsx",
        default="Stimuli.xlsx",
        help="Path to source xlsx file (default: Stimuli.xlsx).",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name. If omitted, the first sheet is used.",
    )
    parser.add_argument(
        "--text-column",
        default="Word",
        help="Header name of the text column to synthesize (default: Word).",
    )
    parser.add_argument(
        "--filename-column",
        default=None,
        help="Optional header name used for output filenames.",
    )
    parser.add_argument(
        "--output-dir",
        default="audio_en",
        help="Directory to write mp3 files (default: audio_en).",
    )
    parser.add_argument(
        "--voice",
        default="en-US-AriaNeural",
        help="Edge TTS voice (default: en-US-AriaNeural).",
    )
    parser.add_argument(
        "--voices",
        default=None,
        help=(
            "Comma-separated voices, e.g. "
            "'male1=en-US-GuyNeural,female1=en-US-AriaNeural'."
        ),
    )
    parser.add_argument(
        "--voice-preset",
        default=None,
        choices=sorted(VOICE_PRESETS.keys()),
        help="Preset voice set.",
    )
    parser.add_argument(
        "--list-presets",
        action="store_true",
        help="Print available voice presets and exit.",
    )
    parser.add_argument(
        "--rate",
        default="+0%",
        help="Speech rate (default: +0%%). Example: -10%%, +15%%.",
    )
    parser.add_argument(
        "--pitch",
        default="+0Hz",
        help="Pitch (default: +0Hz). Example: -20Hz, +20Hz.",
    )
    parser.add_argument(
        "--volume",
        default="+0%",
        help="Volume (default: +0%%).",
    )
    parser.add_argument(
        "--include-row-number",
        action="store_true",
        help="Prefix output filename with the Excel row number.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing mp3 files if present.",
    )
    return parser.parse_args()


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "item"


def voice_alias(value: str) -> str:
    alias = value.replace("Neural", "")
    return slugify(alias)


def parse_voice_specs(voices: str) -> list[tuple[str, str]]:
    specs: list[tuple[str, str]] = []
    for part in voices.split(","):
        raw = part.strip()
        if not raw:
            continue

        if "=" in raw:
            alias, voice = raw.split("=", 1)
            alias = slugify(alias.strip())
            voice = voice.strip()
        else:
            voice = raw
            alias = voice_alias(voice)

        if not alias or not voice:
            raise ValueError(f"Invalid voice spec: '{raw}'")
        specs.append((alias, voice))

    if not specs:
        raise ValueError("No valid voices found in --voices.")
    return specs


def ensure_unique_aliases(specs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    counts: defaultdict[str, int] = defaultdict(int)
    out: list[tuple[str, str]] = []
    for alias, voice in specs:
        counts[alias] += 1
        final_alias = alias if counts[alias] == 1 else f"{alias}_{counts[alias]}"
        out.append((final_alias, voice))
    return out


def resolve_voice_specs(args: argparse.Namespace) -> list[tuple[str, str]]:
    if args.voice_preset and args.voices:
        raise ValueError("Use either --voice-preset or --voices, not both.")
    if args.voice_preset:
        return VOICE_PRESETS[args.voice_preset]
    if args.voices:
        return ensure_unique_aliases(parse_voice_specs(args.voices))
    return [(voice_alias(args.voice), args.voice)]


def get_column_index(headers: list[str], target: str) -> int:
    normalized = {h.strip().lower(): idx for idx, h in enumerate(headers)}
    key = target.strip().lower()
    if key not in normalized:
        raise ValueError(
            f"Column '{target}' not found. Available headers: {', '.join(headers)}"
        )
    return normalized[key]


def load_entries(
    xlsx_path: Path,
    sheet_name: str | None,
    text_column: str,
    filename_column: str | None,
    include_row_number: bool,
) -> list[dict[str, str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError("Sheet is empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    text_idx = get_column_index(headers, text_column)
    file_idx = get_column_index(headers, filename_column) if filename_column else None

    duplicate_counter: defaultdict[str, int] = defaultdict(int)
    entries: list[dict[str, str]] = []

    for row_number, row in enumerate(rows, start=2):
        if row is None:
            continue
        text_value = row[text_idx] if text_idx < len(row) else None
        if text_value is None:
            continue

        text = str(text_value).strip()
        if not text:
            continue

        if file_idx is not None and file_idx < len(row) and row[file_idx] is not None:
            base_source = str(row[file_idx]).strip()
        else:
            base_source = text

        base_name = slugify(base_source)
        duplicate_counter[base_name] += 1
        if duplicate_counter[base_name] > 1:
            base_name = f"{base_name}_{duplicate_counter[base_name]}"

        if include_row_number:
            base_name = f"{row_number:03d}_{base_name}"

        entries.append({"text": text, "filename": f"{base_name}.mp3"})

    wb.close()
    return entries


async def synthesize(
    text: str,
    output_path: Path,
    voice: str,
    rate: str,
    pitch: str,
    volume: str,
) -> None:
    communicate = edge_tts.Communicate(
        text=text,
        voice=voice,
        rate=rate,
        pitch=pitch,
        volume=volume,
    )
    await communicate.save(str(output_path))


async def run(args: argparse.Namespace) -> int:
    voice_specs = resolve_voice_specs(args)
    entries = load_entries(
        xlsx_path=Path(args.xlsx),
        sheet_name=args.sheet,
        text_column=args.text_column,
        filename_column=args.filename_column,
        include_row_number=args.include_row_number,
    )
    if not entries:
        print("No valid text rows found.", file=sys.stderr)
        return 1

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    written = 0
    skipped = 0
    use_voice_subdirs = len(voice_specs) > 1

    for alias, voice in voice_specs:
        voice_output_dir = output_dir / alias if use_voice_subdirs else output_dir
        voice_output_dir.mkdir(parents=True, exist_ok=True)

        for item in entries:
            output_path = voice_output_dir / item["filename"]
            if output_path.exists() and not args.overwrite:
                skipped += 1
                continue

            source_text = item["text"]
            text = PRONUNCIATION_OVERRIDES.get((alias, item["filename"]), source_text)

            await synthesize(
                text=text,
                output_path=output_path,
                voice=voice,
                rate=args.rate,
                pitch=args.pitch,
                volume=args.volume,
            )
            written += 1
            if use_voice_subdirs:
                if text == source_text:
                    print(f"[ok] {alias}/{output_path.name} <- {source_text}")
                else:
                    print(f"[ok] {alias}/{output_path.name} <- {source_text} (override: {text})")
            else:
                if text == source_text:
                    print(f"[ok] {output_path.name} <- {source_text}")
                else:
                    print(f"[ok] {output_path.name} <- {source_text} (override: {text})")

    total = len(entries) * len(voice_specs)
    print(f"Done. written={written}, skipped={skipped}, total={total}")
    return 0


def main() -> int:
    args = parse_args()
    if args.list_presets:
        print("Available presets:")
        for name, specs in VOICE_PRESETS.items():
            print(f"- {name}")
            for alias, voice in specs:
                print(f"  {alias}={voice}")
        return 0
    return asyncio.run(run(args))


if __name__ == "__main__":
    raise SystemExit(main())
