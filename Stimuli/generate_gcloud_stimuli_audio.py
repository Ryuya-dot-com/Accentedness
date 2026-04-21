#!/usr/bin/env python3
"""Generate Accentedness stimulus audio with Google Cloud Text-to-Speech.

This script adapts the UVLT LJT audio pipeline for isolated word stimuli.
It keeps the existing alias directory layout used by the browser task:

    audio_en_6voices/<voice_alias>/<word>.wav

Key differences from the UVLT sentence pipeline:
  - no trailing silence padding is added
  - default sample rate is 44.1 kHz to match the current assets
  - pronunciation overrides are expressed with SSML phoneme tags

Authentication:
  The local ADC quota project should point to `uvlt-cat-ljt`.
  The script validates that before making API calls unless
  `--skip-project-check` is passed.

Examples:
  python generate_gcloud_stimuli_audio.py --dry-run
  python generate_gcloud_stimuli_audio.py --items scapula,acorn
  python generate_gcloud_stimuli_audio.py --output-dir audio_en_6voices_gcloud
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
import sys
import unicodedata
import warnings
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = ROOT / "Stimuli.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "audio_en_6voices"
DEFAULT_META_CSV = DEFAULT_OUTPUT_DIR / "audio_meta.csv"

EXPECTED_PROJECT_ID = "uvlt-cat-ljt"
GENERATOR_VERSION = "gcloud_v1.0-2026-04-21"
TARGET_SAMPLE_RATE = 44100
LUFS_TARGET = -23.0
PEAK_CEILING_DBFS = -1.0
TRIM_TOP_DB = 35
MAX_NORM_ITERS = 3
MIN_DURATION_SEC = 0.20
LOUDNESS_MIN_DURATION_SEC = 0.50

VOICE_PRESETS: dict[str, list[tuple[str, str]]] = {
    "en_us_neural2_3m3f": [
        ("m1_guy", "en-US-Neural2-J"),
        ("m2_christopher", "en-US-Neural2-I"),
        ("m3_ryan", "en-US-Neural2-D"),
        ("f1_aria", "en-US-Neural2-C"),
        ("f2_jenny", "en-US-Neural2-F"),
        ("f3_sonia", "en-US-Neural2-G"),
    ]
}

# Word-level pronunciation fixes for items that TTS often mishandles.
PRONUNCIATION_OVERRIDES: dict[str, dict[str, dict[str, str]]] = {
    "acorn": {
        "default": {
            "alphabet": "ipa",
            "phoneme": "ˈeɪ.kɔɹn",
        }
    },
    "scapula": {
        "default": {
            "alphabet": "ipa",
            "phoneme": "ˈskæp.jə.lə",
        }
    },
}

META_COLUMNS = [
    "audio_file",
    "word",
    "voice_alias",
    "voice_name",
    "input_text",
    "duration_ms",
    "peak_dbfs",
    "integrated_lufs",
    "sha256",
    "sample_rate_hz",
    "project_id",
    "generated_at",
    "generator_version",
]


def import_deps(require: bool = True) -> bool:
    missing = []
    try:
        from google.cloud import texttospeech  # noqa: F401
    except ImportError:
        missing.append("google-cloud-texttospeech")
    try:
        import librosa  # noqa: F401
    except ImportError:
        missing.append("librosa")
    try:
        import numpy  # noqa: F401
    except ImportError:
        missing.append("numpy")
    try:
        import pyloudnorm  # noqa: F401
    except ImportError:
        missing.append("pyloudnorm")
    try:
        import soundfile  # noqa: F401
    except ImportError:
        missing.append("soundfile")

    if missing:
        msg = (
            f"Missing Python dependencies: {', '.join(missing)}\n"
            f"Install with: pip install {' '.join(missing)}"
        )
        if require:
            raise ImportError(msg)
        warnings.warn(msg)
        return False
    return True


def slugify(value: str) -> str:
    text = (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "item"


def parse_voice_specs(voices: str) -> list[tuple[str, str]]:
    specs: list[tuple[str, str]] = []
    for part in voices.split(","):
        raw = part.strip()
        if not raw:
            continue
        if "=" not in raw:
            raise ValueError(
                "Voice specs must be alias=voice_name pairs, "
                f"got: {raw!r}"
            )
        alias, voice = raw.split("=", 1)
        specs.append((slugify(alias), voice.strip()))
    if not specs:
        raise ValueError("No valid voices found in --voices.")
    return specs


def resolve_voice_specs(args: argparse.Namespace) -> list[tuple[str, str]]:
    if args.voice_preset and args.voices:
        raise ValueError("Use either --voice-preset or --voices, not both.")
    if args.voice_preset:
        return VOICE_PRESETS[args.voice_preset]
    if args.voices:
        return parse_voice_specs(args.voices)
    return VOICE_PRESETS["en_us_neural2_3m3f"]


def get_column_index(headers: list[str], target: str) -> int:
    normalized = {h.strip().lower(): idx for idx, h in enumerate(headers)}
    key = target.strip().lower()
    if key not in normalized:
        raise ValueError(
            f"Column '{target}' not found. Available headers: {', '.join(headers)}"
        )
    return normalized[key]


def load_entries(
    xlsx_path: Path,
    sheet_name: str | None,
    text_column: str,
    filename_column: str | None,
    include_row_number: bool,
) -> list[dict[str, str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError("Sheet is empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    text_idx = get_column_index(headers, text_column)
    file_idx = get_column_index(headers, filename_column) if filename_column else None

    duplicate_counter: dict[str, int] = {}
    entries: list[dict[str, str]] = []
    for row_number, row in enumerate(rows, start=2):
        if row is None:
            continue
        text_value = row[text_idx] if text_idx < len(row) else None
        if text_value is None:
            continue

        text = str(text_value).strip()
        if not text:
            continue

        if file_idx is not None and file_idx < len(row) and row[file_idx] is not None:
            base_source = str(row[file_idx]).strip()
        else:
            base_source = text

        stem = slugify(base_source)
        duplicate_counter[stem] = duplicate_counter.get(stem, 0) + 1
        if duplicate_counter[stem] > 1:
            stem = f"{stem}_{duplicate_counter[stem]}"
        if include_row_number:
            stem = f"{row_number:03d}_{stem}"

        entries.append(
            {
                "text": text,
                "stem": stem,
                "filename": f"{stem}.wav",
            }
        )

    wb.close()
    return entries


def resolve_adc_path() -> Path | None:
    env_path = Path.home() / ".config" / "gcloud" / "application_default_credentials.json"
    if env_path.exists():
        return env_path
    return None


def validate_project_id(expected_project_id: str) -> None:
    adc_path = resolve_adc_path()
    if not adc_path:
        raise RuntimeError(
            "ADC credentials not found. Expected ~/.config/gcloud/"
            "application_default_credentials.json"
        )
    data = json.loads(adc_path.read_text(encoding="utf-8"))
    quota_project = data.get("quota_project_id")
    if quota_project != expected_project_id:
        raise RuntimeError(
            "ADC quota project does not match the expected Google Cloud project: "
            f"{quota_project!r} != {expected_project_id!r}"
        )


def build_ssml(text: str, stem: str, voice_alias: str) -> str | None:
    override_spec = PRONUNCIATION_OVERRIDES.get(stem, {})
    override = override_spec.get(voice_alias, override_spec.get("default"))
    if not override:
        return None
    alphabet = override["alphabet"]
    phoneme = override["phoneme"]
    escaped_text = html.escape(text, quote=False)
    escaped_phoneme = html.escape(phoneme, quote=True)
    return (
        "<speak>"
        f'<phoneme alphabet="{alphabet}" ph="{escaped_phoneme}">{escaped_text}</phoneme>'
        "</speak>"
    )


def synthesize_mp3(
    *,
    text: str,
    stem: str,
    voice_alias: str,
    voice_name: str,
    speaking_rate: float,
    pitch: float,
    sample_rate_hz: int,
) -> tuple[bytes, str]:
    from google.cloud import texttospeech

    client = texttospeech.TextToSpeechClient()
    ssml = build_ssml(text, stem, voice_alias)
    if ssml:
        synthesis_input = texttospeech.SynthesisInput(ssml=ssml)
        input_text = ssml
    else:
        synthesis_input = texttospeech.SynthesisInput(text=text)
        input_text = text

    voice = texttospeech.VoiceSelectionParams(
        language_code="en-US",
        name=voice_name,
    )
    audio_config = texttospeech.AudioConfig(
        audio_encoding=texttospeech.AudioEncoding.MP3,
        speaking_rate=speaking_rate,
        pitch=pitch,
        sample_rate_hertz=sample_rate_hz,
    )
    response = client.synthesize_speech(
        input=synthesis_input,
        voice=voice,
        audio_config=audio_config,
    )
    return response.audio_content, input_text


def compute_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def post_process(
    *,
    mp3_bytes: bytes,
    out_wav: Path,
    sample_rate_hz: int,
) -> dict[str, str | int | float]:
    import io

    import librosa
    import numpy as np
    import pyloudnorm as pyln
    import soundfile as sf

    def for_loudness_meter(signal: "np.ndarray", sr_hz: int) -> "np.ndarray":
        min_samples = int(LOUDNESS_MIN_DURATION_SEC * sr_hz)
        if len(signal) >= min_samples:
            return signal
        pad = np.zeros(min_samples - len(signal), dtype=signal.dtype)
        return np.concatenate([signal, pad])

    with io.BytesIO(mp3_bytes) as buf:
        y, sr = librosa.load(buf, sr=sample_rate_hz, mono=True)
    if len(y) == 0:
        raise ValueError("TTS returned empty audio")

    y, _ = librosa.effects.trim(y, top_db=TRIM_TOP_DB)
    if len(y) < int(MIN_DURATION_SEC * sr):
        raise ValueError(
            f"Audio too short after trim ({len(y) / sr:.2f}s): {out_wav.name}"
        )

    peak = float(np.max(np.abs(y)))
    ceiling = 10 ** (PEAK_CEILING_DBFS / 20)
    if peak > ceiling:
        y = y * (ceiling / peak)

    meter = pyln.Meter(sr)
    loudness_in = meter.integrated_loudness(for_loudness_meter(y, sr))
    y = pyln.normalize.loudness(y, loudness_in, LUFS_TARGET)
    for _ in range(MAX_NORM_ITERS):
        peak_after = float(np.max(np.abs(y)))
        if peak_after <= ceiling:
            break
        overshoot_db = 20 * np.log10(peak_after / ceiling)
        y = pyln.normalize.loudness(
            y,
            meter.integrated_loudness(for_loudness_meter(y, sr)),
            LUFS_TARGET - overshoot_db,
        )

    out_wav.parent.mkdir(parents=True, exist_ok=True)
    sf.write(out_wav, y, sr, subtype="PCM_16")

    peak_dbfs = 20 * np.log10(float(np.max(np.abs(y))) + 1e-9)
    duration_ms = int(len(y) / sr * 1000)
    sha256 = compute_sha256(out_wav)
    loudness_final = meter.integrated_loudness(for_loudness_meter(y, sr))
    return {
        "duration_ms": duration_ms,
        "peak_dbfs": round(peak_dbfs, 3),
        "integrated_lufs": round(loudness_final, 3),
        "sha256": sha256,
    }


def load_existing_meta(meta_csv: Path) -> dict[str, dict[str, str]]:
    if not meta_csv.exists():
        return {}
    with meta_csv.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return {row["audio_file"]: row for row in reader}


def write_meta(meta_csv: Path, entries: list[dict[str, str]]) -> None:
    meta_csv.parent.mkdir(parents=True, exist_ok=True)
    entries = sorted(entries, key=lambda row: row["audio_file"])
    with meta_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=META_COLUMNS)
        writer.writeheader()
        writer.writerows(entries)


def process_one(
    *,
    entry: dict[str, str],
    voice_alias: str,
    voice_name: str,
    output_dir: Path,
    project_id: str,
    speaking_rate: float,
    pitch: float,
    sample_rate_hz: int,
    dry_run: bool,
    overwrite: bool,
    existing_meta: dict[str, dict[str, str]],
) -> dict[str, str] | None:
    out_path = output_dir / voice_alias / entry["filename"]
    rel_path = out_path.relative_to(output_dir).as_posix()

    if not overwrite and out_path.exists() and rel_path in existing_meta:
        print(f"  CACHE {rel_path}")
        return existing_meta[rel_path]

    if dry_run:
        print(f"  DRY-RUN {rel_path} <- {entry['text']!r}")
        return None

    print(f"  SYNTH {rel_path} <- {entry['text']!r}")
    mp3_bytes, input_text = synthesize_mp3(
        text=entry["text"],
        stem=entry["stem"],
        voice_alias=voice_alias,
        voice_name=voice_name,
        speaking_rate=speaking_rate,
        pitch=pitch,
        sample_rate_hz=sample_rate_hz,
    )
    audio_meta = post_process(
        mp3_bytes=mp3_bytes,
        out_wav=out_path,
        sample_rate_hz=sample_rate_hz,
    )
    return {
        "audio_file": rel_path,
        "word": entry["stem"],
        "voice_alias": voice_alias,
        "voice_name": voice_name,
        "input_text": input_text,
        "duration_ms": str(audio_meta["duration_ms"]),
        "peak_dbfs": str(audio_meta["peak_dbfs"]),
        "integrated_lufs": str(audio_meta["integrated_lufs"]),
        "sha256": str(audio_meta["sha256"]),
        "sample_rate_hz": str(sample_rate_hz),
        "project_id": project_id,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "generator_version": GENERATOR_VERSION,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate Accentedness word audio via Google Cloud TTS."
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--text-column", default="Word")
    parser.add_argument("--filename-column", default=None)
    parser.add_argument("--include-row-number", action="store_true")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Root output directory for alias subdirectories.",
    )
    parser.add_argument(
        "--meta-csv",
        type=Path,
        default=None,
        help="Metadata CSV path.",
    )
    parser.add_argument(
        "--voice-preset",
        default="en_us_neural2_3m3f",
        choices=sorted(VOICE_PRESETS.keys()),
    )
    parser.add_argument(
        "--voices",
        default=None,
        help="Comma-separated alias=voice_name pairs.",
    )
    parser.add_argument(
        "--items",
        default=None,
        help="Comma-separated word stems to process, e.g. scapula,acorn",
    )
    parser.add_argument(
        "--allow-partial-live-update",
        action="store_true",
        help=(
            "Allow --items updates directly in the live audio_en_6voices tree. "
            "By default, partial refreshes of the live stimulus set are blocked."
        ),
    )
    parser.add_argument("--speaking-rate", type=float, default=1.0)
    parser.add_argument("--pitch", type=float, default=0.0)
    parser.add_argument("--sample-rate-hz", type=int, default=TARGET_SAMPLE_RATE)
    parser.add_argument("--project-id", default=EXPECTED_PROJECT_ID)
    parser.add_argument("--skip-project-check", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def run(args: argparse.Namespace) -> int:
    if not args.xlsx.exists():
        print(f"ERROR: Excel file not found: {args.xlsx}", file=sys.stderr)
        return 2

    meta_csv_path = args.meta_csv or (args.output_dir / "audio_meta.csv")
    live_output_dir = DEFAULT_OUTPUT_DIR.resolve()
    requested_output_dir = args.output_dir.resolve()
    if (
        args.items
        and requested_output_dir == live_output_dir
        and not args.allow_partial_live_update
    ):
        print(
            "ERROR: --items cannot target the live audio_en_6voices directory "
            "unless --allow-partial-live-update is set.",
            file=sys.stderr,
        )
        return 2

    entries = load_entries(
        xlsx_path=args.xlsx,
        sheet_name=args.sheet,
        text_column=args.text_column,
        filename_column=args.filename_column,
        include_row_number=args.include_row_number,
    )
    if args.items:
        wanted = {slugify(item.strip()) for item in args.items.split(",") if item.strip()}
        entries = [entry for entry in entries if entry["stem"] in wanted]
        if not entries:
            print("ERROR: no rows matched --items", file=sys.stderr)
            return 2

    voice_specs = resolve_voice_specs(args)
    existing_meta = load_existing_meta(meta_csv_path)
    all_meta: dict[str, dict[str, str]] = dict(existing_meta)

    if not args.dry_run:
        import_deps(require=True)
        if not args.skip_project_check:
            validate_project_id(args.project_id)

    print(
        f"Processing {len(entries)} words with {len(voice_specs)} voices "
        f"into {args.output_dir}"
    )
    failures: list[tuple[str, str, Exception]] = []
    for voice_alias, voice_name in voice_specs:
        print(f"\n[{voice_alias}] {voice_name}")
        for entry in entries:
            try:
                meta = process_one(
                    entry=entry,
                    voice_alias=voice_alias,
                    voice_name=voice_name,
                    output_dir=args.output_dir,
                    project_id=args.project_id,
                    speaking_rate=args.speaking_rate,
                    pitch=args.pitch,
                    sample_rate_hz=args.sample_rate_hz,
                    dry_run=args.dry_run,
                    overwrite=args.overwrite,
                    existing_meta=existing_meta,
                )
                if meta is not None:
                    all_meta[meta["audio_file"]] = meta
            except Exception as err:  # pragma: no cover - operational path
                print(
                    f"  FAIL {voice_alias}/{entry['stem']}: {err}",
                    file=sys.stderr,
                )
                failures.append((voice_alias, entry["stem"], err))

    if args.dry_run:
        print("\nDry-run complete.")
        return 0

    write_meta(meta_csv_path, list(all_meta.values()))
    print(f"\nMeta file written: {meta_csv_path}")
    if failures:
        print("\nFailures:", file=sys.stderr)
        for voice_alias, stem, err in failures:
            print(f"  {voice_alias}/{stem}: {err}", file=sys.stderr)
        return 1
    return 0


def main() -> None:
    args = parse_args()
    sys.exit(run(args))


if __name__ == "__main__":
    main()
